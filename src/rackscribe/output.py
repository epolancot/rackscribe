import logging
import re
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

log = logging.getLogger("rackscribe")

_INVENTORY_RE = re.compile(
    # Regex used in process_inventory_output to get field values. Implement PID and VID if needed.
    r"""
        NAME:\s*"(?P<name>[^"]+)",\s*
        DESCR:\s*"(?P<description>[^"]+)"\s*
        PID:\s*[^,]*\s*,\s*
        VID:\s*[^,]*\s*,\s*
        SN:\s*(?P<sn>[^\r\n]*)
    """,
    re.MULTILINE | re.VERBOSE,
)


_IOS_PREAMBLE_RE = re.compile(
    # Regex used in remove_config_preamble to strip configuration preamble text in Cisco IOS.
    r"""
    \A
    (?:\s*Building\ configuration\.\.\.\s*\n)?
    (?:\s*Current\ configuration\s*:\s*.*?bytes\s*\n)?
    (?:\n)*
    """,
    re.IGNORECASE | re.VERBOSE,
)


def remove_config_preamble(config: str) -> str:
    """Strip IOS' preamble - Idempotent"""
    # Normalize newlines
    clean_config = config.replace("\r\n", "\n").replace("\r", "\n")
    clean_config = _IOS_PREAMBLE_RE.sub("", clean_config)
    return clean_config.lstrip("\n")


def create_config_file(hostname: str, show_run_output: str) -> None:
    """Write running configuration to a per-host file."""
    log.info(f"Creating configuration file '{hostname}'.")
    path = f"output/configurations/{hostname}.cfg"

    try:
        with open(path, "w") as f:
            f.write(show_run_output)
    except OSError as exc:
        log.error(
            f"Failed to write configuration file for {hostname} at '{path}'. See rackscribe.log for details."
        )
        log.debug(
            f"Failed to write configuration file for {hostname} at '{path}': {exc}", exc_info=True
        )

    log.info(f"File created for {hostname} at '{path}'.")


def process_inventory_output(hostname: str, show_inventory_output: str) -> list[list[str]]:
    rows: list[list[str]] = []
    for m in _INVENTORY_RE.finditer(show_inventory_output):
        name = m.group("name")
        description = m.group("description")
        sn = (m.group("sn") or "").strip() or "N/A"
        rows.append([hostname, name, description, sn])
    return rows


def format_inventory_worksheet(excel_path: str) -> None:
    """Apply basic formatting to the inventory Excel worksheet."""
    wb = load_workbook(excel_path)
    ws = wb.active

    # Freeze header row
    ws.freeze_panes = "A2"

    # Enable auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Header style
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Auto-size columns based on max cell width
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    wb.save(excel_path)


def create_inventory_file(file_name: str, file_path: str, inventory: list[list[str]]) -> None:
    """Write the collected inventory to a timestamped Excel file."""
    now = datetime.now()
    timestamp_str = now.strftime("%Y%m%d-%H%M%S")

    base_name = f"{file_name}_{timestamp_str}"
    full_path = f"{file_path}inventory/{base_name}.xlsx"

    if not inventory:
        log.warning(f"No inventory data to write. Skipping Excel file creation {full_path}.")
        return

    TABLE_COLUMNS = ["Hostname", "Name", "Description", "Serial Number"]

    try:
        df = pd.DataFrame(inventory, columns=TABLE_COLUMNS)
        df.to_excel(f"{full_path}", index=False)
        # Apply formatting AFTER creation
        format_inventory_worksheet(full_path)
    except (OSError, ValueError) as exc:
        log.error(
            f"Failed to write inventory Excel file at '{full_path}'. See rackscribe.log for details."
        )
        log.debug(f"Failed to write inventory Excel file at {full_path}: {exc}", exc_info=True)
        return
    log.info(f"Inventory Excel file succesfully created at {full_path}")
